import openpyxl, pprint

wb = openpyxl.load_workbook('test.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
for i in range(1, sheet.max_row+1):
    for j in range(1, sheet.max_column+1):
        print sheet.cell(row = i, column = j).value
